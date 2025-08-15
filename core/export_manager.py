# -*- coding: utf-8 -*-
"""
Gerenciador de Exportações - Sistema DIRENS
"""

import csv
import os
import logging
from datetime import datetime
from reportlab.lib.pagesizes import A4, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT

class ExportManager:
    """Gerenciador de exportações para CSV e PDF"""
    
    def __init__(self):
        """Inicializa o gerenciador de exportações"""
        self.exports_dir = "exports"
        self.ensure_exports_directory()
    
    def ensure_exports_directory(self):
        """Garante que o diretório de exportações existe"""
        if not os.path.exists(self.exports_dir):
            os.makedirs(self.exports_dir)
    
    def export_csv(self, teachers, school):
        """Exporta professores para CSV"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"professores_{school.replace(' ', '_')}_{timestamp}.csv"
            filepath = os.path.join(self.exports_dir, filename)
            
            with open(filepath, 'w', newline='', encoding='utf-8-sig') as csvfile:
                fieldnames = [
                    'SIAPE', 'Nome', 'CPF', 'Data Nascimento', 'Sexo', 'Estado Civil',
                    'Carga Horária', 'Carreira', 'Data Ingresso', 'Status', 'Área Atuação',
                    'Pós-graduação', 'Graduação', 'Instituição Graduação',
                    'Curso Pós', 'Instituição Pós', 'Email', 'Telefone'
                ]
                
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                for teacher in teachers:
                    writer.writerow({
                        'SIAPE': teacher.get('siape', ''),
                        'Nome': teacher.get('nome', ''),
                        'CPF': teacher.get('cpf', ''),
                        'Data Nascimento': teacher.get('data_nascimento', ''),
                        'Sexo': teacher.get('sexo', ''),
                        'Estado Civil': teacher.get('estado_civil', ''),
                        'Carga Horária': teacher.get('carga_horaria', ''),
                        'Carreira': teacher.get('carreira', ''),
                        'Data Ingresso': teacher.get('data_ingresso', ''),
                        'Status': teacher.get('status', 'Ativo'),
                        'Área Atuação': teacher.get('area_atuacao', ''),
                        'Pós-graduação': teacher.get('pos_graduacao', ''),
                        'Graduação': teacher.get('graduacao', ''),
                        'Instituição Graduação': teacher.get('instituicao_graduacao', ''),
                        'Curso Pós': teacher.get('curso_pos', ''),
                        'Instituição Pós': teacher.get('instituicao_pos', ''),
                        'Email': teacher.get('email', ''),
                        'Telefone': teacher.get('telefone', '')
                    })
            
            logging.info(f"CSV exportado: {filepath}")
            return filepath
            
        except Exception as e:
            logging.error(f"Erro ao exportar CSV: {e}")
            raise
    
    def export_pdf(self, teachers, school):
        """Exporta professores para PDF"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"relatorio_professores_{school.replace(' ', '_')}_{timestamp}.pdf"
            filepath = os.path.join(self.exports_dir, filename)
            
            # Cria documento PDF em orientação paisagem para acomodar mais campos
            from reportlab.lib.pagesizes import landscape
            doc = SimpleDocTemplate(
                filepath,
                pagesize=landscape(A4),
                rightMargin=20,
                leftMargin=20,
                topMargin=30,
                bottomMargin=30
            )
            
            # Estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=TA_CENTER
            )
            
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Heading2'],
                fontSize=12,
                spaceAfter=20,
                alignment=TA_LEFT
            )
            
            # Elementos do documento
            elements = []
            
            # Título
            title = Paragraph(f"RELATÓRIO DE PROFESSORES<br/>{school}", title_style)
            elements.append(title)
            
            # Informações gerais
            info_text = f"""
            <b>Data de Geração:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}<br/>
            <b>Total de Professores:</b> {len(teachers)}<br/>
            <b>Sistema:</b> DIRENS - Controle de Professores v1.0
            """
            
            info_para = Paragraph(info_text, styles['Normal'])
            elements.append(info_para)
            elements.append(Spacer(1, 20))
            
            # Estatísticas resumidas
            stats = self.generate_statistics(teachers)
            if stats:
                elements.append(Paragraph("ESTATÍSTICAS RESUMIDAS", subtitle_style))
                
                stats_data = [
                    ['Categoria', 'Quantidade', 'Percentual'],
                    ['Professores Ativos', str(stats['ativos']), f"{(stats['ativos']/len(teachers)*100):.1f}%" if teachers else "0%"],
                    ['40H Dedicação Exclusiva', str(stats['de_40h']), f"{(stats['de_40h']/len(teachers)*100):.1f}%" if teachers else "0%"],
                    ['Com Doutorado', str(stats['doutorado']), f"{(stats['doutorado']/len(teachers)*100):.1f}%" if teachers else "0%"],
                    ['Com Mestrado', str(stats['mestrado']), f"{(stats['mestrado']/len(teachers)*100):.1f}%" if teachers else "0%"],
                    ['Carreira EBTT', str(stats['ebtt']), f"{(stats['ebtt']/len(teachers)*100):.1f}%" if teachers else "0%"]
                ]
                
                stats_table = Table(stats_data)
                stats_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                elements.append(stats_table)
                elements.append(Spacer(1, 20))
            
            # Lista de professores
            if teachers:
                elements.append(PageBreak())
                elements.append(Paragraph("LISTA COMPLETA DE PROFESSORES", subtitle_style))
                
                # Cabeçalho da tabela com todos os campos
                table_data = [
                    ['SIAPE', 'Nome', 'Nasc.', 'Sexo', 'C.H.', 'Carreira', 'Ingresso', 'Status', 'Área', 'Pós-grad.', 'Graduação', 'Inst.Grad', 'Curso Pós', 'Inst.Pós']
                ]
                
                # Dados dos professores
                for teacher in sorted(teachers, key=lambda x: x.get('nome', '')):
                    table_data.append([
                        teacher.get('siape', '')[:7],  
                        teacher.get('nome', '')[:25],  
                        teacher.get('data_nascimento', '')[:10],
                        teacher.get('sexo', '')[:1],
                        teacher.get('carga_horaria', '')[:6],
                        teacher.get('carreira', '')[:6],
                        teacher.get('data_ingresso', '')[:10],
                        teacher.get('status', 'Ativo')[:8],
                        teacher.get('area_atuacao', '')[:15],
                        teacher.get('pos_graduacao', '')[:10],
                        teacher.get('graduacao', '')[:15],
                        teacher.get('instituicao_graduacao', '')[:12],
                        teacher.get('curso_pos', '')[:12],
                        teacher.get('instituicao_pos', '')[:12]
                    ])
                
                # Cria tabela
                teachers_table = Table(table_data, repeatRows=1)
                teachers_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 7),
                    ('FONTSIZE', (0, 1), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                ]))
                
                elements.append(teachers_table)
            
            # Rodapé
            elements.append(Spacer(1, 30))
            footer_text = f"Relatório gerado pelo Sistema DIRENS - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            footer = Paragraph(footer_text, styles['Normal'])
            elements.append(footer)
            
            # Gera o PDF
            doc.build(elements)
            
            logging.info(f"PDF exportado: {filepath}")
            return filepath
            
        except Exception as e:
            logging.error(f"Erro ao exportar PDF: {e}")
            raise
    
    def export_pdf_with_fields(self, teachers, school, selected_fields):
        """Exporta professores para PDF com campos selecionados"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"relatorio_personalizado_{school.replace(' ', '_')}_{timestamp}.pdf"
            filepath = os.path.join(self.exports_dir, filename)
            
            # Cria documento PDF em orientação paisagem
            from reportlab.lib.pagesizes import landscape
            doc = SimpleDocTemplate(
                filepath,
                pagesize=landscape(A4),
                rightMargin=20,
                leftMargin=20,
                topMargin=30,
                bottomMargin=30
            )
            
            # Estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=TA_CENTER
            )
            
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Heading2'],
                fontSize=12,
                spaceAfter=20,
                alignment=TA_LEFT
            )
            
            # Elementos do documento
            elements = []
            
            # Título
            title = Paragraph(f"RELATÓRIO PERSONALIZADO DE PROFESSORES<br/>{school}", title_style)
            elements.append(title)
            
            # Informações gerais
            info_text = f"""
            <b>Data de Geração:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}<br/>
            <b>Total de Professores:</b> {len(teachers)}<br/>
            <b>Campos Incluídos:</b> {len(selected_fields)}<br/>
            <b>Sistema:</b> DIRENS - Controle de Professores v1.0
            """
            
            info_para = Paragraph(info_text, styles['Normal'])
            elements.append(info_para)
            elements.append(Spacer(1, 20))
            
            # Lista de professores
            if teachers:
                elements.append(Paragraph("LISTA DE PROFESSORES", subtitle_style))
                
                # Cabeçalho da tabela com campos selecionados
                headers = [field['label'] for field in selected_fields]
                table_data = [headers]
                
                # Dados dos professores
                for teacher in sorted(teachers, key=lambda x: x.get('nome', '')):
                    row = []
                    for field in selected_fields:
                        field_key = field['key']
                        value = teacher.get(field_key, '')
                        
                        # Limita tamanho do texto baseado no tipo de campo
                        if field_key == 'nome':
                            value = str(value)[:25]
                        elif field_key in ['data_nascimento', 'data_ingresso']:
                            value = str(value)[:10]
                        elif field_key == 'siape':
                            value = str(value)[:7]
                        elif field_key == 'sexo':
                            value = str(value)[:1]
                        elif field_key in ['carga_horaria', 'carreira', 'status']:
                            value = str(value)[:8]
                        elif field_key in ['pos_graduacao']:
                            value = str(value)[:12]
                        else:
                            value = str(value)[:15]
                        
                        row.append(value)
                    
                    table_data.append(row)
                
                # Calcula larguras das colunas baseado no número de campos
                num_fields = len(selected_fields)
                if num_fields <= 6:
                    font_size_header = 9
                    font_size_data = 8
                elif num_fields <= 10:
                    font_size_header = 8
                    font_size_data = 7
                else:
                    font_size_header = 7
                    font_size_data = 6
                
                # Cria tabela
                teachers_table = Table(table_data, repeatRows=1)
                teachers_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), font_size_header),
                    ('FONTSIZE', (0, 1), (-1, -1), font_size_data),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                ]))
                
                elements.append(teachers_table)
            
            # Rodapé
            elements.append(Spacer(1, 30))
            footer_text = f"Relatório personalizado gerado pelo Sistema DIRENS - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            footer = Paragraph(footer_text, styles['Normal'])
            elements.append(footer)
            
            # Gera o PDF
            doc.build(elements)
            
            logging.info(f"PDF personalizado exportado: {filepath}")
            return filepath
            
        except Exception as e:
            logging.error(f"Erro ao exportar PDF personalizado: {e}")
            raise
    
    def generate_statistics(self, teachers):
        """Gera estatísticas para o relatório"""
        if not teachers:
            return None
        
        total = len(teachers)
        ativos = len([t for t in teachers if t.get('status', 'Ativo') == 'Ativo'])
        de_40h = len([t for t in teachers if t.get('carga_horaria') == '40H_DE'])
        doutorado = len([t for t in teachers if t.get('pos_graduacao') == 'DOUTORADO'])
        mestrado = len([t for t in teachers if t.get('pos_graduacao') == 'MESTRADO'])
        ebtt = len([t for t in teachers if t.get('carreira') == 'EBTT'])
        
        return {
            'total': total,
            'ativos': ativos,
            'de_40h': de_40h,
            'doutorado': doutorado,
            'mestrado': mestrado,
            'ebtt': ebtt
        }
    
    def export_detailed_pdf(self, teachers, school):
        """Exporta relatório PDF detalhado com todos os campos"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"relatorio_detalhado_{school.replace(' ', '_')}_{timestamp}.pdf"
            filepath = os.path.join(self.exports_dir, filename)
            
            doc = SimpleDocTemplate(
                filepath,
                pagesize=letter,
                rightMargin=20,
                leftMargin=20,
                topMargin=30,
                bottomMargin=30
            )
            
            elements = []
            styles = getSampleStyleSheet()
            
            # Título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=14,
                spaceAfter=20,
                alignment=TA_CENTER
            )
            
            title = Paragraph(f"RELATÓRIO DETALHADO DE PROFESSORES - {school}", title_style)
            elements.append(title)
            
            # Para cada professor, criar uma página ou seção
            for i, teacher in enumerate(sorted(teachers, key=lambda x: x.get('nome', ''))):
                if i > 0:
                    elements.append(PageBreak())
                
                # Nome do professor
                name_style = ParagraphStyle(
                    'TeacherName',
                    parent=styles['Heading2'],
                    fontSize=12,
                    spaceAfter=15,
                    alignment=TA_LEFT
                )
                
                name = Paragraph(f"{i+1:03d}. {teacher.get('nome', 'N/A')}", name_style)
                elements.append(name)
                
                # Dados em tabela
                teacher_data = [
                    ['Campo', 'Valor'],
                    ['SIAPE', teacher.get('siape', '')],
                    ['CPF', teacher.get('cpf', '')],
                    ['Data de Nascimento', teacher.get('data_nascimento', '')],
                    ['Sexo', teacher.get('sexo', '')],
                    ['Estado Civil', teacher.get('estado_civil', '')],
                    ['Carga Horária', teacher.get('carga_horaria', '')],
                    ['Carreira', teacher.get('carreira', '')],
                    ['Data de Ingresso', teacher.get('data_ingresso', '')],
                    ['Status', teacher.get('status', 'Ativo')],
                    ['Área de Atuação', teacher.get('area_atuacao', '')],
                    ['Pós-graduação', teacher.get('pos_graduacao', '')],
                    ['Graduação', teacher.get('graduacao', '')],
                    ['Instituição Graduação', teacher.get('instituicao_graduacao', '')],
                    ['Curso Pós-graduação', teacher.get('curso_pos', '')],
                    ['Instituição Pós', teacher.get('instituicao_pos', '')],
                    ['Email', teacher.get('email', '')],
                    ['Telefone', teacher.get('telefone', '')]
                ]
                
                teacher_table = Table(teacher_data, colWidths=[2*inch, 4*inch])
                teacher_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                    ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                elements.append(teacher_table)
                elements.append(Spacer(1, 20))
            
            doc.build(elements)
            
            logging.info(f"PDF detalhado exportado: {filepath}")
            return filepath
            
        except Exception as e:
            logging.error(f"Erro ao exportar PDF detalhado: {e}")
            raise
    
    def export_excel(self, teachers, school):
        """Exporta para Excel usando openpyxl"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"professores_{school.replace(' ', '_')}_{timestamp}.xlsx"
            filepath = os.path.join(self.exports_dir, filename)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Professores"
            
            # Cabeçalhos
            headers = [
                'SIAPE', 'Nome', 'CPF', 'Data Nascimento', 'Sexo', 'Estado Civil',
                'Carga Horária', 'Carreira', 'Data Ingresso', 'Status', 'Área Atuação',
                'Pós-graduação', 'Graduação', 'Instituição Graduação',
                'Curso Pós', 'Instituição Pós', 'Email', 'Telefone'
            ]
            
            # Estilo do cabeçalho
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Adiciona cabeçalhos
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Adiciona dados
            for row, teacher in enumerate(teachers, 2):
                ws.cell(row=row, column=1, value=teacher.get('siape', ''))
                ws.cell(row=row, column=2, value=teacher.get('nome', ''))
                ws.cell(row=row, column=3, value=teacher.get('cpf', ''))
                ws.cell(row=row, column=4, value=teacher.get('data_nascimento', ''))
                ws.cell(row=row, column=5, value=teacher.get('sexo', ''))
                ws.cell(row=row, column=6, value=teacher.get('estado_civil', ''))
                ws.cell(row=row, column=7, value=teacher.get('carga_horaria', ''))
                ws.cell(row=row, column=8, value=teacher.get('carreira', ''))
                ws.cell(row=row, column=9, value=teacher.get('data_ingresso', ''))
                ws.cell(row=row, column=10, value=teacher.get('status', 'Ativo'))
                ws.cell(row=row, column=11, value=teacher.get('area_atuacao', ''))
                ws.cell(row=row, column=12, value=teacher.get('pos_graduacao', ''))
                ws.cell(row=row, column=13, value=teacher.get('graduacao', ''))
                ws.cell(row=row, column=14, value=teacher.get('instituicao_graduacao', ''))
                ws.cell(row=row, column=15, value=teacher.get('curso_pos', ''))
                ws.cell(row=row, column=16, value=teacher.get('instituicao_pos', ''))
                ws.cell(row=row, column=17, value=teacher.get('email', ''))
                ws.cell(row=row, column=18, value=teacher.get('telefone', ''))
            
            # Ajusta largura das colunas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Salva arquivo
            wb.save(filepath)
            
            logging.info(f"Excel exportado: {filepath}")
            return filepath
            
        except ImportError:
            logging.warning("openpyxl não disponível, usando CSV como alternativa")
            return self.export_csv(teachers, school)
        except Exception as e:
            logging.error(f"Erro ao exportar Excel: {e}")
            raise
